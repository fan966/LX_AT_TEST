# -*- coding:utf-8 -*-
import xlrd
import pandas
import os
import json
import openpyxl
from Config.Config import *
import logging
class ExcelUtil(object):
    def __init__(self, filepath, sheet_name = ''):
        """
        ExcelUtil构造方法
        初始化成员变量
        :param filepath: excel文件路径
        :param sheet_name: 指定sheet页名称
        """
        self.filepath = filepath
        self.sheet_name = sheet_name
        self.data = xlrd.open_workbook(self.filepath)
        if sheet_name:
            self.sheet = self.data.sheet_by_name(self.sheet_name)  # 获取sheet页内容
        self.sheetnames = self.data.sheet_names()

    def get_rows(self):
        '''
        获取sheet行数
        :return:
        '''
        rows = self.sheet.nrows
        return rows if rows else None

    def get_cols(self):
        '''
        获取sheet列数
        :return:
        '''
        cols = self.sheet.ncols
        return cols if cols else None

    def read_excel_addlist(self):
        '''
        读取excel数据组合列键值对列表/适合不用转换数据的excel表格
        :return:
        '''
        data_list = []
        rows = self.get_rows()
        cols = self.get_cols()
        rows_value = self.sheet.row_values(0)
        for i in range(1,rows):
            values = self.sheet.row_values(i)
            dic =dict(zip(rows_value,values))
            data_list.append(dic)
        return data_list



    def excel_case_data_info(self, sheet_name_list=[], **rules_info):
        '''
        获取excel接口case数据
        :param sheet_name_list:指定sheet名
        :param r_c_info:过滤规则字典
        :return:
        '''
        result_data = []
        if isinstance(sheet_name_list, list) and sheet_name_list:
            data_dic = pandas.read_excel(self.filepath, sheet_name=sheet_name_list)  # 获取指定sheet数据
        else:
            data_dic = pandas.read_excel(self.filepath, sheet_name=None)  # 获取表格内所有sheet数据
        for i in data_dic: # i 指sheet内所有数据
            query_str = ''
            datas = data_dic[i]
            if rules_info:

                if isinstance(rules_info,dict):
                    for key,value in rules_info.items():
                        if isinstance(value,dict):
                            for key_,value_ in value.items():
                                query_str += str(key_) + ' == "' + str(value_) + '" & '
                            query_str = query_str.rstrip(' &')
                            if not datas.query(query_str).empty:
                                datas1 = datas.query(query_str)
                                result_data += json.loads(datas1.to_json(orient='records')) # Datafrmo格式转换

                        else:
                            logging.error('过滤规则字典的value值必须为字典类型')
                else:
                    logging.error('提供过滤规则字典必须是字段类型')
            else:
                logging.info('过滤规则字段为空，则获取所有测试数据')
                result_data += json.loads(datas.to_json(orient='records'))
        # 循环数据替换数据内指定key的value格式
        for j in result_data:
            for key in j:
                if key in ['NewVerifParmsData','ExpectResult']:
                    j.update({key: json.loads(j[key]) if j[key] else {}})
        return result_data

    def write_result_to_excel(self,row,col,value):
        '''
        写入数据
        :param row:行 openpyxl行列都是从1开始计算
        :param col: 列
        :param data: 写入数据
        :return:
        '''
        try:
            # 获取整个excel数据
            wb = openpyxl.load_workbook(self.filepath)
            # 指定sheet页
            wr = wb[self.sheet_name]
            wr.cell(row, col, value)
            wb.save(self.filepath)
        except Exception as err:
            logging.error(err)
            logging.error('文件写入数据报错，请检查')

    def get_excel_num(self,row_name,row_value):
        '''
        匹配数据在文件行数
        :return:
        '''
        data = self.read_excel_addlist()
        for i in range(len(data)):
            if data[i][row_name] == row_value:
                return i + 2

if __name__ == '__main__':
    path = Config().case_data_file_path
    # ex = ExcelUtil(path,'User')
    # print(ex.get_cols())
    # temp = {'ApplyEnv': '全部', 'IsRun': 'YES', 'ScriptName': 'Api_User_Login_Case.py'}
    # e = ExcelUtil(path)
    # data = e.excel_case_data_info(['User'],t = temp)
    # print(data)
    # e = ExcelUtil(path,'interface')
    # print(e.read_excel_addlist())
    e = ExcelUtil(path)
    dic = {'ApplyEnv': '全部', 'IsRun': 'YES', 'ScriptName': 'Api_Game_GetGameOddsData_Case.py'}
    print(e.excel_case_data_info(sheet_name_list=['Game'],temp=dic))


